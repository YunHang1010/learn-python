from openpyxl import Workbook
import datetime


wb = Workbook()

# grab the active worksheet
ws = wb.active
ws.title = "New Title"

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()

# Loc cells using row and column notation
loc = ws.cell(row=4, column=2, value=10)

# Create 100x100 cells in memory
for x in range(1,101):
    for y in range(1,101):
        ws.cell(row=x, column=y)
        
# Save the file
wb.save("filePath.xlsx")